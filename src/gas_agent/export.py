"""
Export HMISGasRecord list back to Excel by updating the original file.
Preserves all formatting, column widths, styles, etc.
"""

from pathlib import Path
from shutil import copy2

import openpyxl

from gas_agent.schema import HMISGasRecord, COLUMN_INDEX_TO_FIELD


def export_records_to_excel(
    records: list[HMISGasRecord],
    output_path: str | Path,
    *,
    original_path: str | Path,
    sheet_name: str = "Sheet1",
) -> None:
    """
    Update the original Excel file with filled records, preserving formatting.
    
    Args:
        records: List of filled HMISGasRecord objects
        output_path: Where to save the updated file
        original_path: Path to the original Excel file (for formatting)
        sheet_name: Sheet name to update (default: "Sheet1")
    
    Strategy:
        1. Copy original file to output location
        2. Load with openpyxl (preserves formatting)
        3. Update cell values for each record
        4. Save (formatting preserved)
    """
    output_path = Path(output_path)
    original_path = Path(original_path)
    
    # Copy original file to preserve all formatting
    copy2(original_path, output_path)
    
    # Load the copied file (not read_only, not data_only to preserve formatting)
    wb = openpyxl.load_workbook(output_path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    
    # Update data rows (skip header row, start from row 2)
    max_col = max(COLUMN_INDEX_TO_FIELD.keys(), default=0)
    
    for row_idx, record in enumerate(records, start=2):  # Excel rows are 1-indexed, skip header
        record_dict = record.model_dump()
        
        # Update each column for this row
        for col_idx in range(max_col + 1):
            field = COLUMN_INDEX_TO_FIELD.get(col_idx)
            if not field:
                continue
            
            value = record_dict.get(field)
            # Write value to cell (1-indexed column)
            ws.cell(row=row_idx, column=col_idx + 1, value=value)
    
    # Save with formatting preserved
    wb.save(output_path)
    wb.close()
